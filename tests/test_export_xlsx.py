"""Tests for Excel (.xlsx) job export with clickable URL hyperlinks."""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from exports import build_jobs_xlsx  # noqa: E402


def test_build_jobs_xlsx_adds_hyperlinks_for_urls():
    payload = build_jobs_xlsx(
        [
            {
                "Source": "shine",
                "Title": "SEO Analyst",
                "Company": "Acme",
                "Location": "Bangalore",
                "Workplace": "remote",
                "Contract": "Full-time",
                "Published": "3 days ago",
                "Published (ISO)": "2026-07-19T00:00:00+00:00",
                "Salary": "Rs 5 - 8 Lakh/Yr",
                "Job URL": "https://www.shine.com/job/seo/acme/1",
                "Company URL": "https://acme.example",
            },
            {
                "Source": "naukri",
                "Title": "No links",
                "Company": "Beta",
                "Location": "Pune",
                "Workplace": None,
                "Contract": None,
                "Published": None,
                "Published (ISO)": None,
                "Salary": None,
                "Job URL": None,
                "Company URL": "",
            },
        ]
    )
    assert payload[:2] == b"PK"  # zip/xlsx magic

    wb = load_workbook(io.BytesIO(payload))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    job_col = headers.index("Job URL") + 1
    company_col = headers.index("Company URL") + 1

    job_cell = ws.cell(row=2, column=job_col)
    assert job_cell.value == "https://www.shine.com/job/seo/acme/1"
    assert job_cell.hyperlink is not None
    assert job_cell.hyperlink.target == "https://www.shine.com/job/seo/acme/1"

    company_cell = ws.cell(row=2, column=company_col)
    assert company_cell.hyperlink is not None
    assert company_cell.hyperlink.target == "https://acme.example"

    empty_job = ws.cell(row=3, column=job_col)
    assert empty_job.hyperlink is None
